# import module
import openpyxl
from tkinter import *


# load excel with its path
wrkbk = openpyxl.load_workbook("personDetails.xlsx")

sh = wrkbk.active

# iterate through excel and display data
# for i in range(1, sh.max_row+1):
# 	print("\n")
# 	print("Row ", i, " data :")
	
# 	for j in range(1, sh.max_column+1):
# 		cell_obj = sh.cell(row=i, column=j)
# 		print(cell_obj.value, end=" ")

class Table:
     
    def __init__(self,root):
         
        # code for creating table
        for i in range(1, sh.max_row+1):
            for j in range(1, sh.max_column+1):
                 
                self.e = Entry(root, width=10, fg='blue',
                               font=('Arial',16,'bold'))
                 
                self.e.grid(row=i, column=j)
                cell_obj = sh.cell(row=i, column=j)
                data = cell_obj.value
                self.e.insert(END, cell_obj)
 
# take the data
# lst = [(1,'Raj','Mumbai',19),
#        (2,'Aaryan','Pune',18),
#        (3,'Vaishnavi','Mumbai',20),
#        (4,'Rachna','Mumbai',21),
#        (5,'Shubham','Delhi',21)]
  
# find total number of rows and
# columns in list
# total_rows = len(lst)
# total_columns = len(lst[0])
  
# create root window
root = Tk()
t = Table(root)
root.mainloop()