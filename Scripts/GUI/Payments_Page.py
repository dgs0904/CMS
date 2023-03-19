from tkinter import *
from datetime import date, datetime
from openpyxl import load_workbook


window = Tk()
window.geometry("960x540")
window.title("Payments")
window.configure(bg="white")\



# functions to be used.
def num2col(num):
    col = ""

    if(num < 26):
        col += chr(ord('A') + num)
    else:
        col += chr(ord('A') + math.floor(num / 26) - 1)
        col += chr(ord('A') + (num % 26))

    return col

# Main Logic of the page is here.

# -------------------------------GUI Logic.-------------------------------
outlinerFrame = Frame(window, bg="#535c5c", borderwidth=1,relief=SUNKEN)
outlinerFrame.pack(side=LEFT, fill="y")

LEFTFrame = Frame(window, bg="white", border=4)
LEFTFrame.pack(side=LEFT, fill="x")

terminalFrame = Frame(window, bg="white", border=4,)
terminalFrame.pack(side=BOTTOM, fill= "x")

centerFrame = Frame(window, bg="white", border=4,)
centerFrame.pack(fill= "both")

# To display text in outliner.
# outlinerTexr = Label(outlinerFrame, text="Outliner...")
home_button = Button(outlinerFrame, background="black",fg="white",width=20, text="Home")
home_button.pack(pady=10, padx=50)

borrowers_button = Button(outlinerFrame, background="black",fg="white",width=20, text="Borrowers")
borrowers_button.pack(pady=10,padx=50)

loans_button = Button(outlinerFrame, background="black",fg="white",width=20, text="Loans")
loans_button.pack(pady=10,  padx=50)

payments_button = Button(outlinerFrame, background="black",fg="white",width=20, text="Payments")
payments_button.pack(pady=10,  padx=50)


# Widgets to show in center frame.
# Title_label = Label(centerFrame, text="Payements",font="ariel 15 bold",bg= "white", justify= LEFT).grid(row=1, column=3, padx=100, pady =5)
lIDEntry =  Label(centerFrame, text="Loan ID",font="ariel 10 bold",bg= "white", justify= LEFT).grid(row=2, column=1, padx=50, pady =5)
NameEntry =  Label(centerFrame, text="Name",font="ariel 10 bold",bg= "white", justify= LEFT).grid(row=2, column=2, padx=50, pady =5)
EmiEntry =  Label(centerFrame, text="EMI Amount",font="ariel 10 bold",bg= "white", justify= LEFT).grid(row=2, column=3, padx=50, pady =5)
DateEntry=  Label(centerFrame, text="EMI Date",font="ariel 10 bold",bg= "white", justify= LEFT).grid(row=2, column=4, padx=50, pady =5)



wb = load_workbook('./Scripts/loanDetails.xlsx')
ws = wb['Sheet1']
wa = wb['Sheet2']

wp = load_workbook('./Scripts/personDetails.xlsx')
pw = wp['Sheet1']

today = datetime.today()
# label .here
# To find data from particular cells.


Row = 3

for row in range(3,ws.max_row+1):
    # HERE WE WILL TAKE EACH LOAN ID ONE BY ONE FOR EACH ITERATION.
    lID = ws['A' + str(row)]
    # WE WILL TAKE IT'S ROW AS CURRENT ROW TO GO THROUGH EACH EMI CELL STARTING FROM 'O'[NOT ZERO].
    working_row = row
    
    # HERE WE ARE TAKING 14 AS 'O' HAS ASCII VALUE OF 15 AND num2col RETURNS ONE DIGIT GREATER VALUE.
    num = 14
    # THIS VARIABLE IS USED, AS IN WHICH ITERATION pending WRIITEN AND THEN USE THAT ITERATION AS NUMBER OF EMI.
    emiCount = None

    # HERE ws['M' + str(working_row)].value RETURNS THE VALUE STORED IN totalEmi COLUMN THAN TO USE IT AS LIMIT.
    # Lid_Column = 1

    for emi in range(1, int(ws['M' + str(working_row)].value)+1):

        # HERE WE ARE ASKING FOR ALPHABET ONE BY ONE TO START FROM 'O' AS WE ALREADY KNOW THAT EMIs START FROM COLUMN O.
        col = num2col(num)

        # HERE WE ARE CHECKING FROM EMI 1 TO ALL IF THEY HAVE pending WRITTEN IN THEM.
        # IF PENDING IS WRITTEN THEN WE WILL GO IN SHEET2 WHERE DATES FOR EACH EMI IS WRITTEN
        if str(ws[col + str(working_row)].value) == "pending":

            emiCount = emi
            col2 = num2col(num-13)

            if datetime.strptime(str(wa[col2 + str(working_row)].value).split(" ")[0], '%Y-%m-%d') < today:
                # This prints the date of upcoming emi.
                # print(str(wa[col2 + str(working_row)].value).split(" ")[0])
                # lID.value ---> To be Used for fetching the current loan Id.
                loan_ID_value = lID.value
                name_value = pw['B' + str(working_row)].value
                emi_value = ws['H' + str(working_row)].value
                date_value = str(wa[col2 + str(working_row)].value).split(" ")[0]

                lIDEntry =  Label(centerFrame, text=loan_ID_value,font="ariel 10 ",bg= "white", justify= LEFT).grid(row=Row, column=1, padx=40, pady =5)
                NameEntry =  Label(centerFrame, text=name_value,font="ariel 10 ",bg= "white", justify= LEFT).grid(row=Row, column=2, padx=40, pady =5)
                AmountEntry =  Label(centerFrame, text=emi_value,font="ariel 10 ",bg= "white", justify= LEFT).grid(row=Row, column=3, padx=40, pady =5)
                DateEntry =  Label(centerFrame, text=date_value,font="ariel 10 ",bg= "white", justify= LEFT).grid(row=Row, column=4, padx=40, pady =5)
                emiCount

                Row += 1
                # pass
        else:
            print("Something's FishyS")
        num += 1

window.mainloop()