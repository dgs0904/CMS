from openpyxl import workbook, load_workbook
from openpyxl.utils import get_column_letter


# Loading excel file into a object.
wb = load_workbook('personDetails.xlsx')

# Creating active Worksheet from our WorkBook.
ws = wb.active

print(ws)
# Prints the name of worksheet.

print(ws['A1'])
# Prints the Individual Cell.

print(ws['A1'].value)
# Prints the actual cell value.

ws['A2'].value = "Devang"
# To change value of cells.



# tkhtmlview is a module that we use to have html tags in our code.