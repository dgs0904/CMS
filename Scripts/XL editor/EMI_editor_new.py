# Importing Modules 
import math
from datetime import datetime,date
from cProfile import label
from goto import with_goto
from openpyxl import workbook, load_workbook
from openpyxl.utils import get_column_letter

# Runs from A to ZZ only
# A = 0, B = 1...
def num2col(num):
    col = ""

    if(num < 26):
        col += chr(ord('A') + num)
    else:
        col += chr(ord('A') + math.floor(num / 26) - 1)
        col += chr(ord('A') + (num % 26))

    return col

# Declaring Variables. 
lID = None
fName = None
lName = None
numBer = None
aDD = None
aaDnumBer = None
pannumBer = None
bName = None
accnumBer = None
amount = None
comTo = None
comPer = None
comRup = None

working_row = None

myListOfValues = [lID,fName,lName,numBer,aDD,aaDnumBer,pannumBer,bName,accnumBer,amount,comTo,comPer,comRup]



# Main Fucntions.
def emiFinder():
    wb = load_workbook('./Scripts/loanDetails.xlsx')
    ws = wb['Sheet1']
    wa = wb['Sheet2']

    today = datetime.today()
    # label .here
    # To find data from particular cells.
    for row in range(2,ws.max_row+1):
        # HERE WE WILL TAKE EACH LOAN ID ONE BY ONE FOR EACH ITERATION.
        lID = ws['A' + str(row)]
        # WE WILL TAKE IT'S ROW AS CURRENT ROW TO GO THROUGH EACH EMI CELL STARTING FROM 'O'[NOT ZERO].
        working_row = row
        
        # HERE WE ARE TAKING 14 AS 'O' HAS ASCII VALUE OF 15 AND num2col RETURNS ONE DIGIT GREATER VALUE.
        num = 14
        # THIS VARIABLE IS USED, AS IN WHICH ITERATION pending WRIITEN AND THEN USE THAT ITERATION AS NUMBER OF EMI.
        emiCount = None

        # HERE ws['M' + str(working_row)].value RETURNS THE VALUE STORED IN totalEmi COLUMN THAN TO USE IT AS LIMIT.
        for emi in range(1, int(ws['M' + str(working_row)].value)+1):
            # print("Entering loop")
            # HERE WE ARE ASKING FOR ALPHABET ONE BY ONE TO START FROM 'O' AS WE ALREADY KNOW THAT EMIs START FROM COLUMN O.
            col = num2col(num)
            # ws[col + str(working_row)].value = "pending"

            # HERE WE ARE CHECKING FROM EMI 1 TO ALL IF THEY HAVE pending WRITTEN IN THEM.
            if str(ws[col + str(working_row)].value) == "pending":
                # print(f"Line{column}")
                # IF PENDING IS WRITTEN THEN WE WILL GO IN SHEET2 WHERE DATES FOR EACH EMI IS WRITTEN
                emiCount = emi
                #for column in range(1, ws['M' + str(working_row)].value +1):
                # print(f"Line{column}")
                col2 = num2col(num-13)
                # print(col2)
                if datetime.strptime(str(wa[col2 + str(working_row)].value).split(" ")[0], '%Y-%m-%d') < today:
                    print(str(wa[col2 + str(working_row)].value).split(" ")[0])
                    print(f'Loan Id : {lID.value} Emi No. : {emiCount}')
                    # pass
            else:
                print("Something's FishyS")
            num += 1

    
           

    
    wb.save('./Scripts/personDetails.xlsx')
    # return myListOfValues

if __name__ == '__main__':
    emiFinder()