# Importing Modules 
from cProfile import label
from goto import with_goto
from openpyxl import workbook, load_workbook
from openpyxl.utils import get_column_letter


# Declaring Variables. 
ID = None
fName = None
lName = None
numBer = None
aDD = None
aaDnumBer = None
pannumBer = None
bName = None
accnumBer = None

maxLength = 4

myListOfValues = [ID,fName,lName,numBer,aDD,aaDnumBer,pannumBer,bName,accnumBer]



# Main Fucntions.
def dataFinder(ID):
    wb = load_workbook('personDetails.xlsx')
    ws = wb.active


    # To Find how many rows are there.
    for row in range(1,100):
        if ws['A' + str(row)].value != None:
            maxLength = row
           
    #         goto .here

    # print(f'Total rows in the database {maxLength}')

    # label .here
    # To find data from particular cells.
    for row in range(1,4):
        if ID == ws['A' + str(row)].value:
            for col in range(1,10):
                char = get_column_letter(col)
                # print(ws[char+str(row)].value)
                myListOfValues[col - 1] =  ws[char+str(row)].value
                


# To assign data to particular cell.
    # for row in range(2,3):             # Always use 2 for starting Index as 1 is used for User's & Loan's Details.
    #     for col in range(1,14):
    #         char = get_column_letter(col)
    #         ws[char+str(row)] = char + str
    print(myListOfValues)

    print(ws.max_column)
    wb.save('personDetails.xlsx')
    return myListOfValues


if __name__ == '__main__':
    dataFinder('20230227JE')