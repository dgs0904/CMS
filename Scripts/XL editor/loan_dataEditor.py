# Importing Modules 
from datetime import datetime,date,timedelta
from dateutil.relativedelta import relativedelta
from openpyxl import workbook, load_workbook
from openpyxl.utils import get_column_letter

# Runs from A to ZZ only
# A = 0, B = 1...
import math
def num2col(num):
    col = ""

    if(num < 26):
        col += chr(ord('A') + num)
    else:
        col += chr(ord('A') + math.floor(num / 26) - 1)
        col += chr(ord('A') + (num % 26))

    return col

# Declaring Variables.
# HERE WE WILL NOT NEED OTHER DETAILS LIKE NAME, EMAIL OR ID BECAUSE HERE WE'LL OPERATE ON LOAN ID WHICH CREATED ON THE BASIS OF OUR GENERAL ID.
lID = None
amount_taken = None
interestPerc = None
timePeriod = None
payDay = None
startDay = date
endDay = None
emiAmount = None
totalAmount = None
commTo = None
comPer = None
comRup = None
totalEmi= timePeriod

emi_Row = ['A','B','C','D','E','F','G','H','I','J','K','L','M','N','O','R','S','T','U','V','W','X','Y','X']

myListOfValues = [lID,amount_taken,interestPerc,timePeriod,payDay,startDay,endDay,emiAmount,totalAmount,commTo,comPer,comRup,totalEmi]

date_of_emi = None

# Main Fucntions.
def detailWriter(lID,amount_taken,interestPerc,timePeriod,payDay,startDay,endDay,emiAmount,totalAmount,commTo,comPer,comRup,totalEmi):
    wb = load_workbook('./Scripts/loanDetails.xlsx')
    ws = wb["Sheet1"]
    # To write new record about a loan
    ws.append([lID,amount_taken,interestPerc,timePeriod,payDay,startDay,endDay,emiAmount,totalAmount,commTo,comPer,comRup,totalEmi])
    
    # To find the line where the loan was registered and now all EMI blocks will be written out.
    for row in range(1,10):
        # TO FIND THE ROW NUMBER OF LOAN DETAILS
        if lID == ws['A' + str(row)].value:
            working_row = row
        # emi_var = 15
    
    num = 14
    for emi in range(totalEmi):
        col = num2col(num)
        ws[col + str(working_row)].value = "pending"
        num += 1

    num = 1
    print(startDay)
    intialDay = datetime.strptime(str(startDay), '%Y-%m-%d').date()
    wa = wb['Sheet2']
    wa['A' + str(working_row)].value = lID
    for emi in range(0,12):
        col = num2col(num)
        nextDay = intialDay + relativedelta(months = emi)
        wa[col + str(working_row)].value = nextDay
        num += 1

            


    wb.save('./Scripts/loanDetails.xlsx')


if __name__ == '__main__':
    detailWriter("20230225RA2502",5000,0.05,12,"2023-03-09","2023-03-09","2024-03-09",437.5,5250,"Rahul","0.03",1.575,12)