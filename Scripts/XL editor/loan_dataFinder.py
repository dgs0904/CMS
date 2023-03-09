# Importing Modules 
from openpyxl import workbook, load_workbook
from openpyxl.utils import get_column_letter


# Declaring Variables. 
lID = None
fName = None
lName = None
numBer = None
aDD = None
aaDnumBer = None
pannumBer = None
bName = None
accnumBer = None
timePeriod = None
totalEmi= timePeriod
maxLength = 4

myListOfValues = [lID,fName,lName,numBer,aDD,aaDnumBer,pannumBer,bName,accnumBer,totalEmi]



# Main Fucntions.
def dataFinder(lID):
    wb = load_workbook('./Scripts/loanDetails.xlsx')
    ws = wb.active


    # TO FIND TOTAL NUMBERS OF COLUMN IN A SHEETS.
    # print(ws.max_column)

    # label .here
    # To find data from particular cells.
    for row in range(1,4):
        if lID == ws['A' + str(row)].value:
            for col in range(1,11):
                char = get_column_letter(col)
                # print(ws[char+str(row)].value)
                myListOfValues[col - 1] =  ws[char+str(row)].value
                


# To assign data to particular cell.
    # for row in range(2,3):             # Always use 2 for starting Index as 1 is used for User's & Loan's Details.
    #     for col in range(1,14):
    #         char = get_column_letter(col)
    #         ws[char+str(row)] = char + str
    # print(myListOfValues)

    
    wb.save('./Scripts/loanDetails.xlsx')
    return myListOfValues

if __name__ == '__main__':
    dataFinder("20230225RA2502")