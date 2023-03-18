# Importing Modules 
from openpyxl import workbook, load_workbook
from openpyxl.utils import get_column_letter
import sys,os
sys.path.append(os.path.abspath("C:/Users/golu6/Downloads/CMS/Scripts"))

# Declaring Variables.
ID = None
Name = None
email = None
numBer = None
aDD = None
aaDnumBer = None
pannumBer = None
bName = None
accnumBer = None

myListOfValues = [ID,Name,email,numBer,aDD,aaDnumBer,pannumBer,bName,accnumBer]


# Main Fucntions.
def detailWriter(ID,Name,email,numBer,aDD,aaDnumBer,pannumBer,bName,accnumBer):
    wb = load_workbook('./Scripts/personDetails.xlsx')
    ws = wb.active
    # To write new record about a loanee
    ws.append([ID,Name,email,numBer,aDD,aaDnumBer,pannumBer,bName,accnumBer])

    wb.save('./Scripts/personDetails.xlsx')

    print("wrote to workbook")

if __name__ == '__main__':
    detailWriter("20230225RA","Devang Sharma","devang.sharma.7054@gmail.com",8866432894,"Ahmedabad, Sabaramati",12331223123,435554313354,"Apna Bank",124345513)