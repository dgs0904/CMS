from openpyxl import workbook, load_workbook
from openpyxl.utils import get_column_letter


# Creating Variables
lID = None
fName = None
lName = None
numBer = None
aDD = None
aaDnumBer = None
pannumBer = None
bName = None
accnumBer = None
amount = None
comTo = None
comPer = None
comRup = None

maxLength = 2


# Creating list 'myListOfVariable' of variables to store all data in one place.
myListOfValues = [lID,fName,lName,numBer,aDD,aaDnumBer,pannumBer,bName,accnumBer,amount,comTo,comPer,comRup]



# Function to wirte data in Database.
def detailWriter(lID,fName,lName,numBer,aDD,aaDnumBer,pannumBer,bName,accnumBer,amount,comTo,comPer,comRup):
    wb = load_workbook('personDetails.xlsx')
    ws = wb.active

    ws.append([lID,fName,lName,numBer,aDD,aaDnumBer,pannumBer,bName,accnumBer,amount,comTo,comPer,comRup])
    wb.save('personDetails.xlsx')



# Function to find Data in Database.
def dataFinder(lID):
    wb = load_workbook('personDetails.xlsx')
    ws = wb.active

# To find data from particular cells.
    for row in range(1,maxLength):
        if lID == ws['A' + str(row)]:
            for col in range(1,14):
                char = get_column_letter(col)
                myListOfValues[col - 1] =  ws[char+str(row)].value


# To assign data to particular cell.
    # for row in range(2,3):             # Always use 2 for starting Index as 1 is used for Column Names.
    #     for col in range(1,14):
    #         char = get_column_letter(col)
    #         ws[char+str(row)] = char + str

    wb.save('personDetails.xlsx')
    return myListOfValues
    




# To test the fucntion if self run then these values will be passed.
if __name__ == '__main__':
    # detailWriter(1234,'Shaktiman','Purohit',124421124,'Noida',123134134,124315245,'Apna Bank',13526466,38750,'Chandu',.621,4234)
    dataFinder(1234)